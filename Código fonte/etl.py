import openpyxl
from bot import Pje
from time import gmtime, strftime

class Planilha:
    def __init__(self) -> None:
        self._wb = openpyxl.Workbook()
        self._sheet = self._wb.active
        self._nomes_colunas = ['Nº Processo', 'Classe judicial', 'Assunto', 'Jurisdição', 'Autuação', 'Última distribuição', 'Valor da causa', 'Segredo de justiça?', 'Justiça gratuita?', 'Tutela/liminar?', 'Prioridade?', 'Órgão julgador', 'Cargo judicial', 'Competência', 'Polo Ativo (Autor)', 'Polo Ativo (Advogado)', 'Polo Passivo (Autor)', 'Polo Passivo (Advogado)']
        self._processos_falhados = []
        self._contador = 0
        self._qtd_processos_entraram = 0
        self._qtd_processos_sairam = 0
        return None

    def relatorio_falhas(self) -> None:
        with open('Log erros de consulta.txt', 'a') as arquivo:
            momento = strftime('%a, %d %b %Y %H:%M:%S', gmtime())
            arquivo.write(f'{momento}\n')
            for pf in self._processos_falhados:
                arquivo.write(f'{pf}\n')
        return None

    def criar_titulos(self) -> None:
        for col_num, nome_coluna in enumerate(self._nomes_colunas, 1):
            self._sheet.cell(row=1, column=col_num, value=nome_coluna)
        return None
    
    def coletar_dados(self, processos:str) -> None:
        def registrar_dados(dados):
            self._sheet.append(dados[self._contador])
            self._contador += 1
            self._wb.save('resultados.xlsx')
            self._qtd_processos_sairam += 1
            return None
            
        pje = Pje()
        pje.navegar('https://tjrj.pje.jus.br/1g/login.seam')
        pje.autenticar()
        print('Executando...')
        for p in processos:
            self._qtd_processos_entraram += 1
            pje.pagina_pesquisa('https://tjrj.pje.jus.br/1g/Processo/ConsultaProcesso/listView.seam')
            numero_limpo = []
            caracteres_proibidos = ['-', '.', ' ']
            n_processo_cru = str(p)
            if len(p) > 17:
                for numero in p:
                    if numero not in caracteres_proibidos:
                        numero_limpo.append(numero)
                for x in range(0, 3):
                    del numero_limpo[13]
                p = ''
                for l in numero_limpo:
                    p += l
            try:
                dados = pje.pesquisar_processo(p, n_processo_cru)
                registrar_dados(dados)
            except Exception:
                self._processos_falhados.append(n_processo_cru)
                continue
        print(f'Entraram {self._qtd_processos_entraram} processos e saíram {self._qtd_processos_sairam} processos.')
        print('Consultas concluídas!')
        return None
    
    def entradas_saidas(self) -> int:
        entradas = self._qtd_processos_entraram
        saidas = self._qtd_processos_sairam
        return entradas, saidas

def coletar_numeros_processos(caminho_planilha:str) -> list[str]:
    wb = openpyxl.load_workbook(caminho_planilha)
    sheet = wb.active
    primeira_coluna = [cell.value for cell in sheet['A']]
    wb.close()
    return primeira_coluna

def iniciar_bot(caminho_planilha:str) -> tuple[int]:
    plan = Planilha()
    plan.criar_titulos()
    numeros_processos = coletar_numeros_processos(caminho_planilha)
    dados = plan.coletar_dados(numeros_processos)
    entradas_saidas = plan.entradas_saidas()
    plan.relatorio_falhas()
    return entradas_saidas

   